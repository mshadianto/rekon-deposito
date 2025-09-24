"""
Excel Service - Smart Import/Export Handler
File: services/excel_service.py
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, List, Optional, Union
import logging
from pathlib import Path
import tempfile
import os

logger = logging.getLogger(__name__)


class ExcelService:
    """Service for advanced Excel operations"""
    
    def __init__(self):
        self.temp_dir = tempfile.gettempdir()
    
    def read_excel_smart(
        self,
        file_path: str,
        sheet_name: Optional[Union[str, int]] = None,
        skip_rows: int = 0,
        clean_data: bool = True
    ) -> pd.DataFrame:
        """
        Smart Excel reading with automatic cleaning
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name or index
            skip_rows: Number of rows to skip
            clean_data: Whether to clean data
            
        Returns:
            Cleaned DataFrame
        """
        try:
            # Read Excel
            df = pd.read_excel(file_path, sheet_name=sheet_name, skiprows=skip_rows)
            
            if clean_data:
                df = self._clean_dataframe(df)
            
            logger.info(f"Successfully read Excel: {file_path}, Shape: {df.shape}")
            return df
            
        except Exception as e:
            logger.error(f"Error reading Excel {file_path}: {str(e)}")
            raise
    
    def find_header_row(self, file_path: str, sheet_name: str = None) -> int:
        """
        Find the row index where headers are located
        
        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name
            
        Returns:
            Row index of headers
        """
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        # Look for common header keywords
        keywords = ['nomor', 'bilyet', 'rekening', 'nominal', 'tanggal', 'no.', 'tgl']
        
        for idx, row in df.iterrows():
            row_str = ' '.join(str(val).lower() for val in row.values if pd.notna(val))
            if any(keyword in row_str for keyword in keywords):
                logger.info(f"Found header at row {idx}")
                return idx
        
        return 0
    
    def export_to_excel(
        self,
        data_dict: Dict[str, pd.DataFrame],
        output_path: str,
        with_formatting: bool = True
    ) -> bool:
        """
        Export multiple DataFrames to Excel with formatting
        
        Args:
            data_dict: Dict mapping sheet_name to DataFrame
            output_path: Output file path
            with_formatting: Apply formatting
            
        Returns:
            True if successful
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, df in data_dict.items():
                    # Write DataFrame
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    if with_formatting:
                        # Apply formatting
                        worksheet = writer.sheets[sheet_name]
                        self._apply_formatting(worksheet, df)
            
            logger.info(f"Successfully exported to {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            return False
    
    def create_reconciliation_report(
        self,
        summary_df: pd.DataFrame,
        detail_dfs: Dict[str, pd.DataFrame],
        output_path: str
    ) -> bool:
        """
        Create formatted reconciliation report
        
        Args:
            summary_df: Summary DataFrame
            detail_dfs: Dict of detail DataFrames by bank
            output_path: Output file path
            
        Returns:
            True if successful
        """
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Summary sheet
            ws_summary = wb.active
            ws_summary.title = "Summary"
            
            # Write summary with formatting
            self._write_summary_sheet(ws_summary, summary_df)
            
            # Detail sheets
            for bank_name, df in detail_dfs.items():
                ws = wb.create_sheet(title=bank_name[:31])  # Excel limit
                self._write_detail_sheet(ws, df, bank_name)
            
            # Save workbook
            wb.save(output_path)
            logger.info(f"Created reconciliation report: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error creating report: {str(e)}")
            return False
    
    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """Clean DataFrame"""
        
        # Strip whitespace from column names
        df.columns = df.columns.str.strip()
        
        # Remove completely empty rows
        df = df.dropna(how='all')
        
        # Strip whitespace from string columns
        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].str.strip() if df[col].dtype == "object" else df[col]
        
        return df
    
    def _apply_formatting(self, worksheet, df: pd.DataFrame):
        """Apply formatting to worksheet"""
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Format headers
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Auto-adjust column widths
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Apply borders to all cells
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                       min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = border
        
        # Format numeric columns
        for col_idx, col_name in enumerate(df.columns, start=1):
            if pd.api.types.is_numeric_dtype(df[col_name]):
                for row_idx in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.number_format = '#,##0.00'
    
    def _write_summary_sheet(self, worksheet, summary_df: pd.DataFrame):
        """Write summary sheet with formatting"""
        
        # Title
        worksheet['A1'] = "REKONSILIASI DEPOSITO - SUMMARY REPORT"
        worksheet['A1'].font = Font(bold=True, size=14)
        worksheet.merge_cells('A1:F1')
        
        # Write DataFrame starting from row 3
        for r_idx, row in enumerate(dataframe_to_rows(summary_df, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                
                # Format header row
                if r_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        
        # Auto-adjust columns
        for column in worksheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    def _write_detail_sheet(self, worksheet, df: pd.DataFrame, bank_name: str):
        """Write detail sheet with formatting"""
        
        # Title
        worksheet['A1'] = f"Detail Rekonsiliasi - {bank_name}"
        worksheet['A1'].font = Font(bold=True, size=12)
        worksheet.merge_cells('A1:G1')
        
        # Write DataFrame
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 3):
            for c_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                
                if r_idx == 3:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
        
        # Conditional formatting for Status column
        status_col = None
        for idx, col in enumerate(df.columns, 1):
            if 'status' in col.lower():
                status_col = idx
                break
        
        if status_col:
            for row_idx in range(4, worksheet.max_row + 1):
                cell = worksheet.cell(row=row_idx, column=status_col)
                if cell.value == 'Difference':
                    for col_idx in range(1, worksheet.max_column + 1):
                        worksheet.cell(row=row_idx, column=col_idx).fill = PatternFill(
                            start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"
                        )
    
    def merge_excel_files(
        self,
        file_paths: List[str],
        output_path: str,
        sheet_name: str = None
    ) -> bool:
        """
        Merge multiple Excel files into one
        
        Args:
            file_paths: List of Excel file paths
            output_path: Output file path
            sheet_name: Sheet name to read from each file
            
        Returns:
            True if successful
        """
        try:
            dfs = []
            
            for file_path in file_paths:
                df = self.read_excel_smart(file_path, sheet_name=sheet_name)
                dfs.append(df)
            
            # Concatenate
            merged_df = pd.concat(dfs, ignore_index=True)
            
            # Export
            merged_df.to_excel(output_path, index=False)
            
            logger.info(f"Merged {len(file_paths)} files to {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"Error merging files: {str(e)}")
            return False
    
    def validate_excel_structure(
        self,
        file_path: str,
        required_columns: List[str]
    ) -> Dict[str, Union[bool, str]]:
        """
        Validate Excel file structure
        
        Args:
            file_path: Path to Excel file
            required_columns: List of required column names
            
        Returns:
            Validation result dict
        """
        try:
            df = self.read_excel_smart(file_path)
            
            # Check required columns
            missing_cols = [col for col in required_columns if col not in df.columns]
            
            if missing_cols:
                return {
                    'valid': False,
                    'message': f"Missing columns: {', '.join(missing_cols)}",
                    'missing_columns': missing_cols
                }
            
            # Check for empty DataFrame
            if df.empty:
                return {
                    'valid': False,
                    'message': "File is empty",
                    'missing_columns': []
                }
            
            return {
                'valid': True,
                'message': "Validation passed",
                'missing_columns': []
            }
            
        except Exception as e:
            return {
                'valid': False,
                'message': f"Error: {str(e)}",
                'missing_columns': []
            }